"""
routers/admin.py — Panel de administración:
  cuentas, mensualidades, reportes CSV/Excel, correos de alerta
"""
import io
import csv
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from tz import now_col
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from sqlalchemy.orm import Session
from sqlalchemy import func

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL = True
except ImportError:
    OPENPYXL = False

from database import get_db
from models import (
    CuentaUsuario, VehiculoCuenta, QRParqueo,
    SesionParqueo, MensualidadV2, FacturaV2,
)

from auth import requerir_admin, hash_password
from services.qr_service import qr_service
from services.email_service import email_service
from services.invoice_service import invoice_service

router = APIRouter(prefix="/admin", tags=["Administración"])


# ══════════════════════════════════════════════════════════════════
#  SCHEMAS
# ══════════════════════════════════════════════════════════════════
class CuentaAdminCreate(BaseModel):
    nombre:    str
    correo:    EmailStr
    password:  str
    documento: Optional[str] = None
    telefono:  Optional[str] = None
    rol:       str = "usuario"   # "usuario" | "admin"


class MensualidadCreate(BaseModel):
    cuenta_id:    int
    periodo:      str          # "2026-05"
    fecha_inicio: date
    fecha_fin:    date
    monto:        float = 80000.0


class AlertaEmailRequest(BaseModel):
    cuenta_id: int
    asunto:    str
    mensaje:   str


# ══════════════════════════════════════════════════════════════════
#  CUENTAS
# ══════════════════════════════════════════════════════════════════
@router.get("/cuentas")
def listar_cuentas(
    skip: int = 0, limit: int = 100,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    cuentas = db.query(CuentaUsuario).offset(skip).limit(limit).all()
    resultado = []
    for c in cuentas:
        v  = c.vehiculos[0] if c.vehiculos else None
        qr = None
        if v:
            qr = db.query(QRParqueo).filter(
                QRParqueo.vehiculo_id == v.id, QRParqueo.activo == 1
            ).first()
        resultado.append({
            "id":        c.id,
            "nombre":    c.nombre,
            "correo":    c.correo,
            "documento": c.documento,
            "telefono":  c.telefono,
            "rol":       c.rol,
            "activo":    c.activo,
            "vehiculo":  {"placa": v.placa, "marca": v.marca, "tipo_vehiculo": v.tipo_vehiculo} if v else None,
            "tiene_qr":  qr is not None,
            "creado_en": c.creado_en.isoformat(),
        })
    return resultado


@router.post("/cuentas", status_code=201)
def crear_cuenta(
    datos: CuentaAdminCreate,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    correo = datos.correo.lower().strip()
    if db.query(CuentaUsuario).filter(CuentaUsuario.correo == correo).first():
        raise HTTPException(status_code=400, detail="Correo ya registrado")
    if datos.documento:
        if db.query(CuentaUsuario).filter(CuentaUsuario.documento == datos.documento).first():
            raise HTTPException(status_code=400, detail="Documento ya registrado")
    if datos.rol not in ("admin", "usuario"):
        raise HTTPException(status_code=400, detail="Rol inválido")

    c = CuentaUsuario(
        correo=correo,
        nombre=datos.nombre.strip(),
        documento=datos.documento,
        telefono=datos.telefono,
        password_hash=hash_password(datos.password),
        rol=datos.rol,
        activo=1,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"mensaje": "Cuenta creada", "cuenta_id": c.id}


@router.put("/cuentas/{cuenta_id}/toggle")
def toggle_cuenta(
    cuenta_id: int,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    c = db.query(CuentaUsuario).filter(CuentaUsuario.id == cuenta_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Cuenta no encontrada")
    c.activo = 0 if c.activo else 1
    db.commit()
    estado = "activada" if c.activo else "desactivada"
    return {"mensaje": f"Cuenta {estado}", "activo": c.activo}


# ══════════════════════════════════════════════════════════════════
#  MENSUALIDADES
# ══════════════════════════════════════════════════════════════════
@router.get("/mensualidades")
def listar_mensualidades(
    skip: int = 0, limit: int = 100,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    mens = db.query(MensualidadV2).order_by(
        MensualidadV2.creado_en.desc()
    ).offset(skip).limit(limit).all()

    resultado = []
    for m in mens:
        # Contar entradas y salidas (sesiones completadas) dentro del periodo
        entradas = db.query(func.count(SesionParqueo.id)).filter(
            SesionParqueo.cuenta_id == m.cuenta_id,
            SesionParqueo.tipo == "mensualidad",
            SesionParqueo.entrada_at >= m.fecha_inicio,
            SesionParqueo.entrada_at <= m.fecha_fin,
        ).scalar() or 0

        salidas = db.query(func.count(SesionParqueo.id)).filter(
            SesionParqueo.cuenta_id == m.cuenta_id,
            SesionParqueo.tipo == "mensualidad",
            SesionParqueo.estado == "completada",
            SesionParqueo.entrada_at >= m.fecha_inicio,
            SesionParqueo.entrada_at <= m.fecha_fin,
        ).scalar() or 0

        # Tipo de vehículo del usuario
        vehiculo = db.query(VehiculoCuenta).filter(
            VehiculoCuenta.cuenta_id == m.cuenta_id
        ).first()

        resultado.append({
            "id":             m.id,
            "cuenta_id":      m.cuenta_id,
            "usuario":        m.cuenta.nombre,
            "correo":         m.cuenta.correo,
            "periodo":        m.periodo,
            "fecha_inicio":   m.fecha_inicio.isoformat(),
            "fecha_fin":      m.fecha_fin.isoformat(),
            "monto":          float(m.monto),
            "estado":         m.estado,
            "tiene_qr":       m.qr_id is not None,
            "creado_en":      m.creado_en.isoformat(),
            "total_entradas": entradas,
            "total_salidas":  salidas,
            "tipo_vehiculo":  vehiculo.tipo_vehiculo if vehiculo else "carro",
        })
    return resultado


@router.post("/mensualidades", status_code=201)
async def crear_mensualidad(
    datos: MensualidadCreate,
    admin: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    cuenta = db.query(CuentaUsuario).filter(CuentaUsuario.id == datos.cuenta_id).first()
    if not cuenta:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Evitar duplicado en el mismo periodo
    dup = db.query(MensualidadV2).filter(
        MensualidadV2.cuenta_id == datos.cuenta_id,
        MensualidadV2.periodo == datos.periodo,
    ).first()
    if dup:
        raise HTTPException(
            status_code=400,
            detail=f"Ya existe mensualidad para {datos.periodo}"
        )

    # Generar factura de mensualidad
    total = db.query(func.count(FacturaV2.id)).scalar() or 0
    num   = f"FAC-{date.today().year}-{total + 1:05d}"
    fac   = FacturaV2(
        numero=num,
        cuenta_id=datos.cuenta_id,
        tipo="mensualidad",
        monto=Decimal(str(datos.monto)),
        descripcion=f"Mensualidad parqueadero | Periodo: {datos.periodo}",
        metodo_pago="efectivo",
        estado="pagada",
    )
    db.add(fac)
    db.flush()

    mens = MensualidadV2(
        cuenta_id=datos.cuenta_id,
        periodo=datos.periodo,
        fecha_inicio=datos.fecha_inicio,
        fecha_fin=datos.fecha_fin,
        monto=Decimal(str(datos.monto)),
        estado="activa",
        factura_id=fac.id,
    )
    db.add(mens)
    db.commit()
    db.refresh(mens)

    # Enviar notificación
    try:
        await email_service.enviar_mensualidad(
            correo=cuenta.correo,
            nombre=cuenta.nombre,
            numero=num,
            periodo=datos.periodo,
            fecha_inicio=datos.fecha_inicio,
            fecha_fin=datos.fecha_fin,
            monto=datos.monto,
        )
    except Exception:
        pass

    return {"mensaje": "Mensualidad creada", "mensualidad_id": mens.id}


@router.post("/mensualidades/{mens_id}/generar-qr")
def generar_qr_mensualidad(
    mens_id: int,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    mens = db.query(MensualidadV2).filter(MensualidadV2.id == mens_id).first()
    if not mens:
        raise HTTPException(status_code=404, detail="Mensualidad no encontrada")

    cuenta = mens.cuenta

    # Desactivar QR de mensualidad anteriores para esta cuenta
    db.query(QRParqueo).filter(
        QRParqueo.cuenta_id == cuenta.id,
        QRParqueo.tipo == "mensualidad",
        QRParqueo.activo == 1,
    ).update({"activo": 0})

    codigo  = qr_service.generar_codigo_unico(cuenta.id) + "_M"
    img_b64 = qr_service.generar_imagen_qr(codigo, cuenta.nombre)

    qr = QRParqueo(
        cuenta_id=cuenta.id,
        codigo=codigo,
        tipo="mensualidad",
        activo=1,
        imagen_b64=img_b64,
    )
    db.add(qr)
    db.flush()
    mens.qr_id = qr.id
    db.commit()

    return {
        "mensaje":    "QR de mensualidad generado",
        "codigo":     codigo,
        "imagen_b64": img_b64,
    }


@router.put("/mensualidades/{mens_id}/estado")
def cambiar_estado_mensualidad(
    mens_id: int,
    datos: dict,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    mens = db.query(MensualidadV2).filter(MensualidadV2.id == mens_id).first()
    if not mens:
        raise HTTPException(status_code=404, detail="Mensualidad no encontrada")
    nuevo = datos.get("estado", "pendiente")
    if nuevo not in ("activa", "vencida", "pendiente", "cancelada"):
        raise HTTPException(status_code=400, detail="Estado inválido")
    mens.estado = nuevo
    db.commit()
    return {"mensaje": f"Estado actualizado a '{nuevo}'"}


@router.delete("/mensualidades/{mens_id}")
def eliminar_mensualidad(
    mens_id: int,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    mens = db.query(MensualidadV2).filter(MensualidadV2.id == mens_id).first()
    if not mens:
        raise HTTPException(status_code=404, detail="Mensualidad no encontrada")
    mens.estado = "cancelada"
    db.commit()
    return {"mensaje": "Mensualidad cancelada"}


# ══════════════════════════════════════════════════════════════════
#  DASHBOARD ADMIN
# ══════════════════════════════════════════════════════════════════
@router.get("/dashboard")
def dashboard(
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    hoy = date.today()

    sesiones_hoy = db.query(func.count(SesionParqueo.id)).filter(
        func.date(SesionParqueo.entrada_at) == hoy
    ).scalar() or 0

    activas = db.query(func.count(SesionParqueo.id)).filter(
        SesionParqueo.estado == "activa"
    ).scalar() or 0

    recaudo_hoy = db.query(func.sum(FacturaV2.monto)).filter(
        func.date(FacturaV2.fecha_emision) == hoy,
        FacturaV2.estado == "pagada",
        FacturaV2.tipo == "visita",
    ).scalar() or 0

    mensualidades_activas = db.query(func.count(MensualidadV2.id)).filter(
        MensualidadV2.estado == "activa",
        MensualidadV2.fecha_inicio <= hoy,
        MensualidadV2.fecha_fin >= hoy,
    ).scalar() or 0

    total_usuarios = db.query(func.count(CuentaUsuario.id)).filter(
        CuentaUsuario.rol == "usuario"
    ).scalar() or 0

    return {
        "sesiones_hoy":         sesiones_hoy,
        "vehiculos_activos":    activas,
        "recaudo_hoy":          float(recaudo_hoy),
        "mensualidades_activas": mensualidades_activas,
        "total_usuarios":       total_usuarios,
        "fecha":                hoy.isoformat(),
    }


# ══════════════════════════════════════════════════════════════════
#  REPORTES CSV / EXCEL
# ══════════════════════════════════════════════════════════════════
def _sesiones_query(db: Session, tipo_reporte: str, fecha_inicio=None, fecha_fin=None):
    q = db.query(SesionParqueo)
    if tipo_reporte == "visitas":
        q = q.filter(SesionParqueo.tipo.in_(["visita", "camara"]))
    elif tipo_reporte == "mensualidades":
        q = q.filter(SesionParqueo.tipo == "mensualidad")
    if fecha_inicio:
        q = q.filter(SesionParqueo.entrada_at >= fecha_inicio)
    if fecha_fin:
        q = q.filter(SesionParqueo.entrada_at <= fecha_fin)
    return q.order_by(SesionParqueo.entrada_at.desc()).all()


@router.get("/reportes/visitas")
def reporte_visitas(
    formato: str = Query("csv", enum=["csv", "excel"]),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d") if fecha_inicio else None
    ff = datetime.strptime(fecha_fin,    "%Y-%m-%d") if fecha_fin    else None
    sesiones = _sesiones_query(db, "visitas", fi, ff)

    headers_cols = [
        "ID", "Placa", "Usuario", "Tipo", "Entrada", "Salida",
        "Duración (min)", "Valor ($)", "Estado", "Factura"
    ]

    def fila(s):
        fac = s.factura.numero if s.factura else ""
        return [
            s.id, s.placa,
            s.cuenta.nombre if s.cuenta else "visitante",
            s.tipo,
            s.entrada_at.strftime("%Y-%m-%d %H:%M") if s.entrada_at else "",
            s.salida_at.strftime("%Y-%m-%d %H:%M") if s.salida_at else "",
            s.duracion_minutos or "",
            float(s.valor_cobrado) if s.valor_cobrado is not None else "",
            s.estado, fac,
        ]

    if formato == "excel" and OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Visitas"

        # Encabezado
        header_fill = PatternFill("solid", fgColor="1A2744")
        for col, h in enumerate(headers_cols, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, s in enumerate(sesiones, 2):
            for col, val in enumerate(fila(s), 1):
                ws.cell(row=row, column=col, value=val)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nombre = f"reporte_visitas_{date.today()}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre}"},
        )

    # CSV fallback
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers_cols)
    for s in sesiones:
        writer.writerow(fila(s))
    output.seek(0)
    nombre = f"reporte_visitas_{date.today()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@router.get("/reportes/mensualidades")
def reporte_mensualidades(
    formato: str = Query("csv", enum=["csv", "excel"]),
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    mens = db.query(MensualidadV2).order_by(MensualidadV2.creado_en.desc()).all()

    headers_cols = [
        "ID", "Usuario", "Correo", "Periodo",
        "Inicio", "Fin", "Monto ($)", "Estado", "Factura"
    ]

    def fila(m):
        fac = m.factura.numero if m.factura else ""
        return [
            m.id, m.cuenta.nombre, m.cuenta.correo, m.periodo,
            m.fecha_inicio.isoformat(), m.fecha_fin.isoformat(),
            float(m.monto), m.estado, fac,
        ]

    if formato == "excel" and OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Mensualidades"

        header_fill = PatternFill("solid", fgColor="1A2744")
        for col, h in enumerate(headers_cols, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, m in enumerate(mens, 2):
            for col, val in enumerate(fila(m), 1):
                ws.cell(row=row, column=col, value=val)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nombre = f"reporte_mensualidades_{date.today()}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre}"},
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers_cols)
    for m in mens:
        writer.writerow(fila(m))
    output.seek(0)
    nombre = f"reporte_mensualidades_{date.today()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


# ══════════════════════════════════════════════════════════════════
#  EMAIL DE ALERTA
# ══════════════════════════════════════════════════════════════════
@router.post("/email/alerta")
async def enviar_alerta(
    datos: AlertaEmailRequest,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    cuenta = db.query(CuentaUsuario).filter(CuentaUsuario.id == datos.cuenta_id).first()
    if not cuenta:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    enviado = await email_service.enviar_alerta(
        correo=cuenta.correo,
        nombre=cuenta.nombre,
        asunto=datos.asunto,
        mensaje=datos.mensaje,
    )
    return {
        "enviado": enviado,
        "correo":  cuenta.correo,
        "mensaje": "Alerta enviada correctamente" if enviado else "Email no configurado (logged en consola)",
    }


# ══════════════════════════════════════════════════════════════════
#  FACTURAS — descargar PDF / reenviar email
# ══════════════════════════════════════════════════════════════════
@router.get("/facturas/{factura_id}/pdf")
def descargar_factura_pdf(
    factura_id: int,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    """Descarga la factura como PDF (admin)."""
    factura = db.query(FacturaV2).filter(FacturaV2.id == factura_id).first()
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    cuenta = db.query(CuentaUsuario).filter(CuentaUsuario.id == factura.cuenta_id).first()

    pdf = _generar_pdf_factura(factura, cuenta, db)
    if not pdf:
        raise HTTPException(status_code=500, detail="fpdf2 no está instalado o error al generar PDF")

    return StreamingResponse(
        iter([pdf]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Factura_{factura.numero}.pdf"'},
    )


@router.post("/facturas/{factura_id}/reenviar")
async def reenviar_factura(
    factura_id: int,
    _: CuentaUsuario = Depends(requerir_admin),
    db: Session = Depends(get_db),
):
    """Reenvía la factura por email al usuario (admin)."""
    factura = db.query(FacturaV2).filter(FacturaV2.id == factura_id).first()
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    cuenta = db.query(CuentaUsuario).filter(CuentaUsuario.id == factura.cuenta_id).first()
    if not cuenta:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    pdf = _generar_pdf_factura(factura, cuenta, db)

    if factura.tipo == "visita":
        sesion = db.query(SesionParqueo).filter(
            SesionParqueo.factura_id == factura.id
        ).first()
        enviado = await email_service.enviar_factura(
            correo=cuenta.correo,
            nombre=cuenta.nombre,
            doc_cliente=cuenta.documento or "",
            numero=factura.numero,
            placa=sesion.placa if sesion else "N/A",
            entrada=sesion.entrada_at if sesion else now_col(),
            salida=sesion.salida_at if sesion else now_col(),
            duracion=f"{sesion.duracion_minutos} min" if sesion else "N/A",
            minutos=sesion.duracion_minutos if sesion else 0,
            monto=float(factura.monto),
            pdf_bytes=pdf,
        )
    else:
        mens = db.query(MensualidadV2).filter(
            MensualidadV2.factura_id == factura.id
        ).first()
        enviado = await email_service.enviar_mensualidad(
            correo=cuenta.correo,
            nombre=cuenta.nombre,
            doc_cliente=cuenta.documento or "",
            numero=factura.numero,
            periodo=mens.periodo if mens else "N/A",
            fecha_inicio=str(mens.fecha_inicio) if mens else "N/A",
            fecha_fin=str(mens.fecha_fin) if mens else "N/A",
            monto=float(factura.monto),
            pdf_bytes=pdf,
        )

    return {
        "enviado": enviado,
        "correo":  cuenta.correo,
        "factura": factura.numero,
        "mensaje": f"Factura reenviada a {cuenta.correo}" if enviado else "SMTP no configurado",
    }


# ── Helper: generar PDF según tipo ───────────────────────────────
def _generar_pdf_factura(factura: FacturaV2, cuenta, db: Session) -> Optional[bytes]:
    if not cuenta:
        return None
    if factura.tipo == "visita":
        sesion = db.query(SesionParqueo).filter(
            SesionParqueo.factura_id == factura.id
        ).first()
        return invoice_service.generar_visita(
            numero=factura.numero,
            nombre_cliente=cuenta.nombre,
            doc_cliente=cuenta.documento or "",
            correo_cliente=cuenta.correo,
            placa=sesion.placa if sesion else "N/A",
            entrada=sesion.entrada_at if sesion else factura.fecha_emision,
            salida=sesion.salida_at if sesion else factura.fecha_emision,
            duracion_str=f"{sesion.duracion_minutos} min" if sesion else "N/A",
            minutos=sesion.duracion_minutos if sesion else 0,
            valor=float(factura.monto),
        )
    else:
        mens = db.query(MensualidadV2).filter(
            MensualidadV2.factura_id == factura.id
        ).first()
        return invoice_service.generar_mensualidad(
            numero=factura.numero,
            nombre_cliente=cuenta.nombre,
            doc_cliente=cuenta.documento or "",
            correo_cliente=cuenta.correo,
            periodo=mens.periodo if mens else "N/A",
            fecha_inicio=str(mens.fecha_inicio) if mens else "N/A",
            fecha_fin=str(mens.fecha_fin) if mens else "N/A",
            monto=float(factura.monto),
        )
